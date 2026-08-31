"""Reader for the Bath (Connaught Mansions) Tinytag workbook export.

Architecture Layer 1. Turns the four quarterly ``.xlsx`` workbooks into the unified long schema.

Three properties of this export drive the implementation:

1. **Temperature and humidity share a row and timestamp.** Unlike LaSDPC — where T and RH came
   from separate devices logging seconds apart, so an exact pivot paired only 5 of 109,081
   readings (DD-006) — the Tinytag loggers write both channels together at a 5-minute cadence.
   Pairing is therefore native and ``schema.resample_long`` is *not* needed for this source.
2. **Sheet names are inconsistent across the four workbooks.** The same physical sensor appears
   as ``erht2_outsidefront``, ``erht2 external front``, and ``erht2_ExternalFront``; the hall is
   ``rht19_hall`` in one file and ``rht19 hallway`` in another. Loading naively yields 11
   "sensors" instead of 9, silently splitting two rooms in half.
3. **Column E is an Excel formula string**, not a value (``openpyxl`` in read-only mode returns
   the formula text). Only columns A-C are read.

Gaps are preserved, never interpolated: ~16 days are missing between the first two workbooks
(2024-01-10 to 2024-01-26), plus shorter gaps in March and May. The hall sensor starts
2024-01-26 and the external-rear sensor 2024-03-19.
"""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd

from dissertation_code import config
from dissertation_code.data import schema

logger = logging.getLogger(__name__)

#: Canonical sensor names, keyed by the workbook sheet name normalised to lowercase alphanumerics.
#: Resolves the cross-workbook naming drift described in the module docstring.
SENSOR_ALIASES = {
    "erht2outsidefront": "external_front",
    "erht2externalfront": "external_front",
    "ehrt1externalrear": "external_rear",
    "rht19hall": "hall",
    "rht19hallway": "hall",
    "rht11livingroom": "livingroom",
    "rht12bathroom": "bathroom",
    "rht13study": "study",
    "rht17kitchen": "kitchen",
    "rht18spareroom": "spareroom",
    "rht7bedroom": "bedroom",
}

#: Sensors outside the dwelling. Excluded from comfort modelling (nobody occupies them), but
#: used to derive the outdoor running mean that drives clothing insulation (see comfort/clothing).
EXTERNAL_SENSORS = ("external_front", "external_rear")

#: Worksheet name fragment marking the derived per-day min/avg/max summaries, which are not source
#: data and are skipped.
_DAILY_SUFFIX = "daily"

#: Zero-based column positions in each data sheet: timestamp, temperature, humidity.
#: Column D (logger dew point) is redundant with T+RH; column E is an Excel formula string.
_COL_TIMESTAMP = 0
_COL_TEMPERATURE = 1
_COL_HUMIDITY = 2


def normalise_sheet_name(name: str) -> str:
    """Reduce a workbook sheet name to its alias key (lowercase, no spaces or underscores)."""
    return name.lower().replace(" ", "").replace("_", "")


def canonical_sensor(sheet_name: str) -> str | None:
    """Map a sheet name to its canonical sensor name, or None if the sheet is not source data.

    Returns None for chart sheets and the derived ``*_Daily`` summary sheets.
    """
    key = normalise_sheet_name(sheet_name)
    if key.endswith(_DAILY_SUFFIX):
        return None
    return SENSOR_ALIASES.get(key)


def read_workbook(path: Path) -> pd.DataFrame:
    """Read one quarterly workbook into a wide frame (timestamp, zone, temperature, humidity).

    Rows with a missing timestamp or a non-numeric reading are dropped — the export is clean in
    practice (zero bad rows across all 514k readings), but the guard keeps a malformed cell from
    silently becoming NaN downstream.
    """
    # openpyxl is an optional-at-import dependency: only the Bath path needs it.
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    records: list[tuple] = []
    try:
        for sheet_name in workbook.sheetnames:
            zone = canonical_sensor(sheet_name)
            if zone is None:
                continue
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
                timestamp = row[_COL_TIMESTAMP]
                temperature = row[_COL_TEMPERATURE]
                humidity = row[_COL_HUMIDITY]
                if timestamp is None:
                    continue
                if not isinstance(temperature, (int, float)):
                    continue
                if not isinstance(humidity, (int, float)):
                    continue
                records.append((timestamp, zone, float(temperature), float(humidity)))
    finally:
        workbook.close()

    return pd.DataFrame(records, columns=list(schema.WIDE_COLUMNS))


def load_wide(
    data_dir: Path = config.BATH_DATASET_DIR,
    workbooks: tuple[str, ...] = config.BATH_WORKBOOKS,
    include_external: bool = False,
) -> pd.DataFrame:
    """Load all Bath workbooks into one wide frame, deduplicated and chronologically sorted.

    Args:
        data_dir: directory holding the quarterly workbooks.
        workbooks: workbook filenames, in chronological order.
        include_external: keep the two outdoor sensors. False for comfort modelling (nobody
            occupies the outdoors); True when deriving the outdoor running mean for clothing.

    Returns:
        A frame with schema.WIDE_COLUMNS: timestamp, zone, temperature, relative_humidity.
    """
    frames = []
    for filename in workbooks:
        path = Path(data_dir) / filename
        if not path.exists():
            raise FileNotFoundError(f"Bath workbook not found: {path}")
        frame = read_workbook(path)
        logger.info("Read %s: %d readings", filename, len(frame))
        frames.append(frame)

    wide = pd.concat(frames, ignore_index=True)
    wide[schema.TIMESTAMP] = pd.to_datetime(wide[schema.TIMESTAMP])

    if not include_external:
        wide = wide[~wide[schema.ZONE].isin(EXTERNAL_SENSORS)]

    # Workbook date ranges abut rather than overlap, but a shared boundary reading would appear
    # twice; drop exact (zone, timestamp) duplicates so the pool has no repeated instances.
    wide = wide.drop_duplicates(subset=[schema.ZONE, schema.TIMESTAMP])

    wide = wide.sort_values([schema.ZONE, schema.TIMESTAMP]).reset_index(drop=True)
    logger.info(
        "Bath dataset: %d readings across %d sensors (%s to %s)",
        len(wide),
        wide[schema.ZONE].nunique(),
        wide[schema.TIMESTAMP].min(),
        wide[schema.TIMESTAMP].max(),
    )
    return wide


def to_long(wide: pd.DataFrame) -> pd.DataFrame:
    """Melt a Bath wide frame into the canonical long schema.

    The long form is the adapter contract shared with ``load_lasdpc`` — downstream code targets
    one shape regardless of source. That Bath happens to arrive already paired is an internal
    detail of this module.
    """
    long = wide.melt(
        id_vars=[schema.TIMESTAMP, schema.ZONE],
        value_vars=[schema.TEMPERATURE, schema.RELATIVE_HUMIDITY],
        var_name=schema.CHANNEL,
        value_name=schema.VALUE,
    )[list(schema.LONG_COLUMNS)]
    return schema.validate_long(long.reset_index(drop=True))
